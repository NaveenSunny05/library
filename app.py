import os
from datetime import datetime
from flask import Flask, render_template, request, send_file
import openpyxl
from openpyxl.styles import Font

app = Flask(__name__)

# Configure directories
DATASET_FOLDER = 'datasets'
OUTPUT_FOLDER = 'outputs'

app.config['DATASET_FOLDER'] = DATASET_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER

# Ensure directories exist
os.makedirs(DATASET_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


def find_roll_number(roll_number):
    """Search for roll number in dataset Excel files."""
    student_details = None
    for filename in os.listdir(DATASET_FOLDER):
        if filename.endswith(('.xlsx', '.xls')):
            filepath = os.path.join(DATASET_FOLDER, filename)
            wb = openpyxl.load_workbook(filepath)

            for sheet in wb.worksheets:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if str(row[0]) == str(roll_number):
                        student_details = {
                            'name': row[1],
                            'roll_number': row[0],
                            'branch': row[2] if len(row) > 2 else 'N/A'
                        }
                        break
                if student_details:
                    break
            if student_details:
                break
    return student_details


def count_present_students():
    """Count students without 'Time Out' in Digital Library for today."""
    count = 0
    today = datetime.now().strftime("%d-%m-%y")
    today_file = f'Attendance{today}.xlsx'
    today_path = os.path.join(OUTPUT_FOLDER, today_file)

    if os.path.exists(today_path):
        wb = openpyxl.load_workbook(today_path)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[5] is None or row[5] == '':
                    count += 1
    return count


def update_attendance_sheet(student_details, entry_time, sheet_name='Attendance', is_dl=False):
    """Update attendance sheet for the student."""
    today = datetime.now().strftime("%d-%m-%y")
    output_folder = OUTPUT_FOLDER
    output_file = os.path.join(output_folder, f'{sheet_name}_{today}.xlsx')

    # Create workbook if file doesn't exist
    if not os.path.exists(output_file):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = sheet_name
        sheet.append(['S.No', 'Name', 'Roll Number', 'Branch', 'Time In', 'Time Out'])
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        wb.save(output_file)

    # Load workbook and update sheet
    wb = openpyxl.load_workbook(output_file)
    sheet = wb[sheet_name]

    # Check for existing entry without 'Time Out'
    existing_entry_index = None
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if str(row[2]) == str(student_details['roll_number']) and not row[5]:
            existing_entry_index = idx
            break

    if existing_entry_index:
        # Update 'Time Out'
        sheet.cell(row=existing_entry_index, column=6, value=entry_time)
    else:
        # Append new entry
        next_sno = len(list(sheet.rows))
        sheet.append([
            next_sno,
            student_details['name'],
            student_details['roll_number'],
            student_details['branch'],
            entry_time,
            None
        ])

    wb.save(output_file)


@app.route('/', methods=['GET', 'POST'])
def index():
    message = None
    student_details = None
    roll_number_submitted = False

    if request.method == 'POST':
        roll_number = request.form.get('roll_number')
        action = request.form.get('action')

        # Step 1: Roll number entered
        if roll_number and not action:
            student_details = find_roll_number(roll_number)
            if student_details:
                roll_number_submitted = True
            else:
                message = "Roll number not found"

        # Step 2: Action selected
        elif action and roll_number:
            student_details = find_roll_number(roll_number)
            if student_details:
                current_time = datetime.now().strftime("%H:%M:%S")

                # Update attendance
                update_attendance_sheet(student_details, current_time, 'Attendance')
                message = "Library Attendance recorded successfully"
            else:
                message = "Roll number not found"

    present_count = count_present_students()
    return render_template(
        'index.html',
        message=message,
        student_details=student_details,
        roll_number_submitted=roll_number_submitted,
        present_count=present_count
    )


@app.route('/download_attendance/<filename>')
def download_attendance(filename):
    folder = OUTPUT_FOLDER 
    return send_file(os.path.join(folder, filename), as_attachment=True)


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')

